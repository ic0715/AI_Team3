"""Build personas_with_goals.jsonl from the CliftonStrengths Excel.

Source: docs/ralph_loop/CliftonStrengths_*.xlsx (located under the repo)

Sheets used:
  - "페르소나 전체"        (36x23, rows 3-36 are the 34 personas)
  - "Goal"                  (103x7, 34 personas x 3 goals)
  - "시뮬레이션 주의사항"   (36x1, rows 2-36, persona-aligned text notes)

Output:
  - ralph_loop/data/personas_with_goals.jsonl  (one JSON object per line, persona_id 1..34)
  - ralph_loop/data/atypical_personas.txt      (7 IDs, one per line)

Goal-selection rule (docs/ralph_loop/ralph_loop_spec.md §4.1):
  1. Among the 3 goal rows for the persona, pick rows where specificity_level == "medium".
  2. If exactly one matches, use it.
  3. If zero or 2+, fall back to the row with the smallest original sheet index.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from .config import DATA_DIR, excel_path

OUT = DATA_DIR / "personas_with_goals.jsonl"
ATYP_OUT = DATA_DIR / "atypical_personas.txt"

PERSONA_SHEET = "페르소나 전체"
GOAL_SHEET = "Goal"
NOTE_SHEET = "시뮬레이션 주의사항"

# Persona sheet columns (row 2 = header, rows 3..36 = 34 personas)
PERSONA_COLS = [
    ("persona_id", 1),
    ("nickname", 2),
    ("birthdate", 3),
    ("gender", 4),
    ("current_job_field", 5),
    ("career_years", 6),
    ("current_concern", 7),
    ("strength_1", 8),
    ("strength_2", 9),
    ("strength_3", 10),
    ("strength_4", 11),
    ("strength_5", 12),
    ("company_size", 13),
    ("industry", 14),
    ("specific_role", 15),
    ("domain_distribution", 16),
    ("primary_domain", 17),
    ("personality", 18),
    ("strength_combo_core", 19),
    ("trigger_type", 20),
    ("trigger_event", 21),
    ("three_months_before", 22),
    ("remarks", 23),
]

ATYPICAL_NICKNAMES = {
    "박서연", "이채린", "김하은", "강현수", "한가람", "김다은", "조시현",
}


def _cell(ws, r, c):
    v = ws.cell(r, c).value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def load_personas(wb):
    ws = wb[PERSONA_SHEET]
    personas = {}
    for r in range(3, ws.max_row + 1):  # rows 3.. (34-persona sheet ends at 36; small sheets end earlier)
        pid = _cell(ws, r, 1)
        if pid is None:
            continue
        rec = {name: _cell(ws, r, col) for name, col in PERSONA_COLS}
        rec["top5_strengths"] = [
            rec.pop(f"strength_{i}") for i in range(1, 6)
        ]
        personas[int(pid)] = rec
    return personas


def load_goals(wb):
    """Return {persona_id: [goal_dict, ...]} in sheet order.

    Returns {} when the workbook has no Goal sheet (e.g. the 2-persona
    add-on file). In that case the caller supplies goals via --goals JSON.
    """
    if GOAL_SHEET not in wb.sheetnames:
        return {}
    ws = wb[GOAL_SHEET]
    goals = {}
    for r in range(2, ws.max_row + 1):
        pid = _cell(ws, r, 1)
        if pid is None:
            continue
        pid = int(pid)
        entry = {
            "axis_code": _cell(ws, r, 3),
            "goal": _cell(ws, r, 4),
            "secret_goal": _cell(ws, r, 5),
            "specificity_level": _cell(ws, r, 6),
            "emotional_tone": _cell(ws, r, 7),
            "_row": r,
        }
        goals.setdefault(pid, []).append(entry)
    return goals


def load_notes(wb):
    if NOTE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[NOTE_SHEET]
    notes = {}
    # rows 2..36, persona_id assumed aligned with persona sheet order (1..34)
    for i, r in enumerate(range(2, 36), start=1):
        v = _cell(ws, r, 1)
        if v:
            notes[i] = v
    return notes


def pick_goal(goal_list):
    """ralph_loop_spec.md §4.1 deterministic pick: medium-first, else smallest index."""
    medium = [g for g in goal_list if (g.get("specificity_level") or "").lower() == "medium"]
    if len(medium) == 1:
        chosen = medium[0]
    else:
        chosen = min(goal_list, key=lambda g: g["_row"])
    chosen = dict(chosen)
    chosen.pop("_row", None)
    return chosen


def _load_external_goals(path: Path) -> dict:
    """Load a goals JSON keyed by persona_id (string or int).

    Each value is a single goal dict (axis_code/goal/secret_goal/
    specificity_level/emotional_tone). Used when the Excel has no Goal sheet.
    Keys starting with '_' (e.g. '_comment') are ignored.
    """
    raw = json.loads(path.read_text(encoding="utf-8"))
    out = {}
    for k, v in raw.items():
        if str(k).startswith("_"):
            continue
        out[int(k)] = v
    return out


def main(argv=None):
    import argparse

    ap = argparse.ArgumentParser(description="Build personas_with_goals JSONL from a CareerPT persona Excel.")
    ap.add_argument("--excel", type=str, default=None,
                    help="Path to the persona Excel. Default: newest CliftonStrengths_*.xlsx under docs/ralph_loop/.")
    ap.add_argument("--goals", type=str, default=None,
                    help="Path to an external goals JSON (keyed by persona_id). Used when the Excel has no Goal sheet.")
    ap.add_argument("--out", type=str, default=None,
                    help=f"Output JSONL path. Default: {OUT}")
    ap.add_argument("--append", action="store_true",
                    help="Append to the output file instead of overwriting (keeps existing personas).")
    args = ap.parse_args(argv)

    xlsx = Path(args.excel) if args.excel else excel_path()
    out_path = Path(args.out) if args.out else OUT
    external_goals = _load_external_goals(Path(args.goals)) if args.goals else {}

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    personas = load_personas(wb)
    goals = load_goals(wb)
    notes = load_notes(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    atypical_ids = []
    mode = "a" if args.append else "w"
    written = 0
    with out_path.open(mode, encoding="utf-8") as f:
        for pid in sorted(personas):
            rec = personas[pid]
            gl = goals.get(pid, [])
            if gl:
                chosen = pick_goal(gl)
                all_goals = [{k: v for k, v in g.items() if k != "_row"} for g in gl]
            elif pid in external_goals:
                chosen = dict(external_goals[pid])
                all_goals = [chosen]
            else:
                print(f"WARN: persona {pid} ({rec.get('nickname')}) has no goal entries (sheet or --goals) — skipped")
                continue
            rec["selected_goal"] = chosen
            rec["all_goals"] = all_goals
            rec["simulation_note"] = notes.get(pid)
            rec["is_atypical"] = rec["nickname"] in ATYPICAL_NICKNAMES
            if rec["is_atypical"]:
                atypical_ids.append(pid)
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            written += 1

    # Only rewrite the atypical-id index for the canonical 34-persona build
    # (the add-on personas are not atypical and must not clobber the index).
    if not args.append and atypical_ids:
        ATYP_OUT.write_text("\n".join(str(p) for p in atypical_ids) + "\n", encoding="utf-8")
        print(f"Atypical persona_ids: {atypical_ids} → {ATYP_OUT}")
    print(f"Wrote {written} personas → {out_path} (mode={mode})")


if __name__ == "__main__":
    main()
